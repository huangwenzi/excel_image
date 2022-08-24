from ast import If
from unittest import case
from openpyxl import Workbook
from openpyxl import load_workbook



import help as Help
import excel_lib as ExcelLib

# 行宽、高 于图片像素比例 100*100 == 800*130




# 26个字母  
chr_list = Help.get_26_char()
# 读取配置
cfg = Help.get_cfg()
# 读取图片列表
file_list = ExcelLib.get_dir_image_list()
sheet_max_num = cfg['最多多少行'] * cfg['每行多少列']
    
# 创建excel
excel = Workbook()
# 创建总览页
first_sheet_name = '总览'
sheet_index = 0
first_sheet = ExcelLib.create_sheet(excel, first_sheet_name, sheet_index)
image_idx = 1
for tmp_image_obj in file_list:
    ExcelLib.sheet_add_image(first_sheet, image_idx, tmp_image_obj.path)
    image_idx += 1
# 创建分页
sheet_index += 1
next_sheet = ExcelLib.create_sheet(excel, str(sheet_index), sheet_index)
image_idx = 1
sheet_name = ""
for tmp_image_obj in file_list:
    sheet_name = sheet_name + tmp_image_obj.name + "、"
    for item in range(tmp_image_obj.num):
        ExcelLib.sheet_add_image(next_sheet, image_idx, tmp_image_obj.path)
        image_idx += 1
        # 本页是否已满
        if image_idx == sheet_max_num:
            image_idx = 1
            # 修改本页名
            next_sheet.title = sheet_name[:-1]
            # 创建新页
            sheet_index += 1
            next_sheet = ExcelLib.create_sheet(excel, str(sheet_index), sheet_index)
            # 重置页名
            sheet_name = tmp_image_obj.name + "、"
# 重命名最后一页
if image_idx != 1:
    next_sheet.title = sheet_name[:-1]

# 保存excel
excel.save('./out.xlsx')



































